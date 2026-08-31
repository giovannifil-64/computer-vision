"""
make_excel
==========
Collect every result of the as-is evaluation into a single spreadsheet: the
headline metrics, the stage-by-stage diagnostic, and the per-subject tables the
aggregates are computed from. The workbook content is written in Italian, since
it is meant to be shared with the project's collaborators.

Functions
---------
- `sheet(wb, title, rows, widths, note)`: Write one formatted sheet.
- `main()`: Build the workbook from the CSV and JSON produced by the pipeline.

Example
-------
```bash
python make_excel.py --output ../Output_prepost --report ../report_prepost
```

Notes
-----
- Reads only files already produced by the evaluation, so it can be re-run at any
  time without recomputing anything.
"""
import os
import csv
import json
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEAD_FILL = PatternFill('solid', fgColor='E8EEF7')
HEAD_FONT = Font(bold=True)

IT_COLS = {
    'subject': 'soggetto', 'n_teeth': 'denti valutati',
    'skipped_extracted': 'esclusi: estratti', 'skipped_mismatch': 'esclusi: rimeshati',
    'rot_err': 'errore rotazione (gradi)', 'rot_baseline': 'baseline rotazione (gradi)',
    'trans_err': 'errore traslazione', 'trans_baseline': 'baseline traslazione',
    'pcd_err': 'errore point cloud (mm)', 'pcd_baseline': 'baseline point cloud (mm)',
}


def sheet(wb, title, rows, widths=None, note=None, header=True):
    """
    Write one sheet with a bold header row and sensible column widths.

    Parameters
    ----------
    - `wb (Workbook)`: Target workbook.
    - `title (str)`: Sheet name.
    - `rows (list)`: List of row lists; the first is the header when `header` is True.
    - `widths (list, optional)`: Column widths.
    - `note (str, optional)`: Explanatory line written under the table.
    - `header (bool, optional)`: Whether the first row is a header. Default `True`.

    Returns
    -------
    - `Worksheet`: The sheet just written.
    """
    ws = wb.create_sheet(title)
    for r in rows:
        ws.append(r)
    if header and rows:
        for c in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HEAD_FONT
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.freeze_panes = 'A2'
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if note:
        ws.append([])
        ws.append([note])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)
    return ws


def main():
    """Build `clik_as_is_results.xlsx` from the evaluation outputs."""
    ap = argparse.ArgumentParser(description="Collect the as-is results into one spreadsheet.")
    ap.add_argument('--output', required=True, help='folder with alignment_metrics.csv')
    ap.add_argument('--report', required=True, help='where to write the workbook')
    ap.add_argument('--data', default=None,
                    help='folder with the diagnostic JSON files (default: <report>/data)')
    args = ap.parse_args()
    args.data = args.data or os.path.join(args.report, 'data')

    wb = Workbook()
    wb.remove(wb.active)

    # ---- headline metrics -------------------------------------------------
    sheet(wb, 'Riepilogo', [
        ['Metrica (per dente, media +/- dev.std)', 'CLIK as-is', 'Nessun movimento', 'Paper (detector ri-addestrato)'],
        ['Rotazione (gradi)', '11.00 +/- 8.56', '11.05 +/- 10.27', '6.69 +/- 2.56'],
        ['Traslazione', '2.18 +/- 1.56', '2.01 +/- 1.79', '1.20 +/- 0.44'],
        ['Point cloud (mm)', '3.10 +/- 2.18', '2.15 +/- 2.06', '1.30 +/- 0.68'],
        [],
        ['Mediane e tasso di vittoria', 'CLIK', 'Nessun movimento', 'CLIK migliore su'],
        ['Rotazione (gradi)', 8.99, 8.59, '48% dei denti'],
        ['Traslazione', 1.81, 1.58, '43% dei denti'],
        ['Point cloud (mm)', 2.47, 1.53, '24% dei denti'],
        [],
        ['Soggetti valutati', 246],
        ['Denti valutati', 6223],
        ['Denti esclusi - estratti durante il trattamento', 310],
        ['Denti esclusi - rimeshati tra i due stadi', 93],
    ], widths=[46, 18, 20, 30],
        note='"Nessun movimento" = errore che si otterrebbe lasciando ogni dente dove si trova; e\' il riferimento che una predizione utile deve battere. '
             'Test as-is: pesi pre-addestrati, nessun fine-tuning, modalita\' crown-only (le scansioni intraorali non hanno le radici).')

    # ---- stage decomposition (read from the diagnostics, never hardcoded) ----
    def _load(name):
        p = os.path.join(args.data, name)
        return json.load(open(p)) if os.path.exists(p) else None

    rep, acc_o, acc_f, dif = (_load('landmark_repeatability.json'), _load('landmark_accuracy.json'),
                              _load('landmark_accuracy_final.json'), _load('diffusion_error.json'))
    rows = [['Stadio', 'Cosa misura', 'Risultato', 'Campione']]
    if rep:
        n = len(rep['per_subject'])
        rows.append(['1 - rilevamento landmark (ripetibilita\')',
                     'landmark trovati sul post vs quelli sul pre trasportati con la trasformazione vera',
                     f"{rep['pose_median_mm']:.2f} mm mediana", f'{n} soggetti'])
        rows.append(['1 - rumore di campionamento (controllo)',
                     'stessa mesh rilevata due volte con seed diversi',
                     f"{rep['noise_median_mm']:.2f} mm mediana", f'{n} soggetti'])
    for tag, a in (('pre-trattamento', acc_o), ('post-trattamento', acc_f)):
        if a:
            skill = 1 - a['overall_median_mm'] / a['baseline_mm']
            rows.append([f'1 - accuratezza anatomica ({tag})',
                         'distanza tra ogni landmark annotato nel dataset e il landmark CLIK piu\' vicino',
                         f"{a['overall_median_mm']:.2f} mm vs {a['baseline_mm']:.2f} mm del baseline casuale (skill {skill:.2f})",
                         'annotazioni del dataset'])
    if dif:
        n = len(dif['per_subject'])
        rows.append(['2 - diffusion',
                     'landmark target predetti vs quelli ideali (landmark iniziali sotto la trasformazione vera)',
                     f"{dif['diffusion_median_mm']:.2f} mm mediana", f'{n} soggetti'])
        rows.append(['3 - fit rigido',
                     'parte della nuvola di landmark predetta che non e\' un movimento rigido',
                     f"{dif['rigid_residual_median_mm']:.2f} mm mediana", f'{n} soggetti'])
    sheet(wb, 'Diagnostica per stadio', rows, widths=[42, 70, 54, 26],
          note='Lo Stage 1 e\' sia ripetibile (sotto il rumore di campionamento) sia anatomicamente corretto, e lo Stage 3 perde quasi nulla: '
               'l\'errore si concentra nella diffusion. I valori sono letti dai file di diagnostica, quindi seguono sempre l\'ultimo run.')

    # ---- landmark accuracy per class -------------------------------------
    acc_path = os.path.join(args.data, 'landmark_accuracy.json')
    if os.path.exists(acc_path):
        acc = json.load(open(acc_path))
        base = acc['baseline_mm']
        rows = [['Classe di landmark', 'CLIK (mm)', 'Baseline casuale (mm)', 'Skill (1 - CLIK/baseline)']]
        for k in sorted(acc['per_class']):
            v = acc['per_class'][k]
            rows.append([k, round(v, 3), round(base, 3), round(1 - v / base, 3)])
        rows.append(['TOTALE', round(acc['overall_median_mm'], 3), round(base, 3),
                     round(1 - acc['overall_median_mm'] / base, 3)])
        sheet(wb, 'Accuratezza landmark', rows, widths=[22, 14, 24, 26],
              note='Skill: 0 = come il caso, 1 = perfetto. E\' una misura di copertura: i due schemi di landmark sono diversi, '
                   'quindi mostra che CLIK mette un landmark su ogni punto annotato, non una corrispondenza per classe.')

    # ---- extra analyses ---------------------------------------------------
    detail_path = os.path.join(args.data, 'metrics_detail.json')
    if os.path.exists(detail_path):
        ex = json.load(open(detail_path))

        col = ex.get('collision', {})
        if col:
            lab = {'initial': 'Pre-trattamento', 'pred': 'Predizione CLIK', 'final': 'Ground truth (post reale)'}
            sheet(wb, 'Collisione', [['Dentatura', 'Profondita\' compenetrazione (mm)', 'Punti compenetrati (%)', 'Soggetti']] +
                  [[lab.get(k, k), round(v['median_depth_mm'], 4), round(100 * v['penetrating_fraction'], 2), v['n']]
                   for k, v in col.items()],
                  widths=[28, 32, 26, 12],
                  note='Quarta metrica del paper. Un minimo di compenetrazione e\' fisiologica (il trattamento porta i denti a contatto stretto), '
                       'ma la predizione ne produce circa tre volte la profondita\' del setup clinico reale e coinvolge quattro volte piu\' punti.')

        sig = ex.get('significance', {})
        if sig:
            sheet(wb, 'Significativita', [['Metrica', 'CLIK (media)', 'Baseline (media)', 'p (t-test appaiato)', 'p (Wilcoxon)', 'CLIK migliore?']] +
                  [[k, round(v['clik_mean'], 3), round(v['baseline_mean'], 3),
                    f"{v['p_ttest']:.2e}", f"{v['p_wilcoxon']:.2e}", 'si' if v['clik_better'] else 'no']
                   for k, v in sig.items()],
                  widths=[16, 16, 18, 22, 18, 16],
                  note='Test appaiati sui soggetti: la differenza rispetto al riferimento "nessun movimento" e\' significativa su tutte e tre le metriche.')

        tt = ex.get('by_tooth_type', {})
        if tt:
            sheet(wb, 'Per tipo di dente', [['Tipo', 'Denti', 'Rotazione CLIK (gradi)', 'Rotazione baseline', 'Point cloud CLIK (mm)', 'Point cloud baseline']] +
                  [[k, v['n'], round(v['rot_median'], 3), round(v['rot_baseline'], 3),
                    round(v['pcd_median'], 3), round(v['pcd_baseline'], 3)] for k, v in tt.items()],
                  widths=[14, 10, 22, 20, 22, 22],
                  note='In assoluto i molari sono i piu\' accurati, ma sono anche quelli che si muovono meno: rispetto al baseline sono i peggiori. '
                       'CLIK regge meglio in senso relativo sui denti anteriori, cioe\' quelli che il trattamento sposta di piu\'.')

        exc = ex.get('excluded', {})
        if exc:
            disp = exc.get('displacement_applied_to_extracted_teeth_mm') or {}
            sheet(wb, 'Denti esclusi', [['Categoria', 'Denti', 'Nota']] +
                  [['Terzi molari', exc.get('third_molars_never_predicted', 0),
                    'lo schema di CLIK copre 28 denti: non vengono mai predetti'],
                   ['Estratti durante il trattamento', exc.get('teeth_extracted_during_treatment', 0),
                    f"assenti nel post; CLIK non lo sa e li sposta comunque di {round(disp.get('median', 0), 2)} mm mediani"],
                   ['Rimeshati fra i due stadi', 93, 'niente corrispondenza vertice-a-vertice, esclusi dalle metriche']],
                  widths=[34, 10, 78])

    # ---- seed variability -------------------------------------------------
    sv_path = os.path.join(args.data, 'seed_variability.json')
    if os.path.exists(sv_path):
        sv = json.load(open(sv_path))
        rows = [['Run (seed)', 'Point cloud mediana (mm)']]
        rows += [[os.path.basename(r), round(m, 4)] for r, m in zip(sv['runs'], sv['pcd_median_per_run'])]
        rows += [['baseline (nessun movimento)', round(sv['pcd_baseline_median'], 4)], []]
        rows += [['Dispersione fra seed (dev.std, mm)', round(sv['pcd_std_mm'], 4)],
                 ['Dispersione fra seed (rotazione, gradi)', round(sv['rot_std_deg'], 4)],
                 ['Soggetti', sv['n_subjects']]]
        sheet(wb, 'Variabilita seed', rows, widths=[38, 26],
              note='La diffusion e\' generativa, quindi ogni esecuzione campiona diversamente. Il divario dal baseline e\' circa 11 volte la dispersione fra seed: '
                   'l\'errore misurato e\' sistematico e non rumore, e valutare con un solo seed era affidabile.')

    # ---- per-subject alignment -------------------------------------------
    csv_path = os.path.join(args.output, 'alignment_metrics.csv')
    if os.path.exists(csv_path):
        rd = list(csv.reader(open(csv_path)))
        head = [IT_COLS.get(h, h) for h in rd[0]]
        rows = [head] + [[r[0]] + [int(x) if i < 3 else round(float(x), 4)
                                   for i, x in enumerate(r[1:])] for r in rd[1:]]
        sheet(wb, 'Per soggetto - allineamento', rows, widths=[12] + [18] * (len(head) - 1),
              note='Una riga per soggetto. Le colonne "baseline" sono il riferimento "nessun movimento" calcolato sullo stesso soggetto.')

    # ---- per-subject diagnostics -----------------------------------------
    rep_path = os.path.join(args.data, 'landmark_repeatability.json')
    if os.path.exists(rep_path):
        rep = json.load(open(rep_path))['per_subject']
        sheet(wb, 'Per soggetto - Stage 1',
              [['soggetto', 'landmark', 'ripetibilita\' (mm)', 'rumore di fondo (mm)']] +
              [[r['sid'], r['n'], round(r['pose_median'], 4),
                round(r['noise_median'], 4) if r['noise_median'] is not None else ''] for r in rep],
              widths=[12, 12, 22, 24],
              note='La ripetibilita\' e\' mediamente inferiore al rumore di campionamento: il detector e\' stabile rispetto alla posa.')

    dif_path = os.path.join(args.data, 'diffusion_error.json')
    if os.path.exists(dif_path):
        dif = json.load(open(dif_path))['per_subject']
        sheet(wb, 'Per soggetto - Stage 2',
              [['soggetto', 'errore diffusion (mm)', 'residuo non rigido (mm)']] +
              [[r['sid'], round(r['diff_median_mm'], 4), round(r['rigid_residual_mm'], 4)] for r in dif],
              widths=[12, 24, 26],
              note='Il residuo non rigido e\' molto minore dell\'errore: la nuvola predetta e\' quasi un movimento rigido dell\'input, ma quello sbagliato.')

    out = os.path.join(args.report, 'clik_as_is_results.xlsx')
    wb.save(out)
    print(f'scritto {out}  ({len(wb.sheetnames)} fogli: {", ".join(wb.sheetnames)})')


if __name__ == '__main__':
    main()
